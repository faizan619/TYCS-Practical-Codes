import openpyxl
  
wb = openpyxl.Workbook()  
sheet = wb.active
heading=["Studentname","Sub1","Sub2","Sub3","Total"]
names=["Carls","James","Paul","Philip","Smith","Thomson","Rhodey","Stark","Gary","AnneMarie"]
marks=[50,45,60,55,70,45,67,78,89,90]
row_count=1

for col in range(1,len(heading)+1):
    sheet.cell(row=row_count,column=col).value=heading[col-1]
row_count+=1

for i in range(len(marks)):
    sheet.cell(row=row_count,column=1).value=names[i]
    sheet.cell(row=row_count,column=2).value=marks[i]
    sheet.cell(row=row_count,column=3).value=marks[i]
    sheet.cell(row=row_count,column=4).value=marks[i]
    sheet.cell(row=row_count,column=5).value=marks[i]*3
    row_count+=1
wb.save("students.xlsx")
